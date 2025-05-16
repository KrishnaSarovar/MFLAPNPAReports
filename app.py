import pandas as pd 
import streamlit as st
import numpy as np
import openpyxl
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, numbers
import re
import os
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from copy import copy
import plotly.express as px


# Ask user about the report name
report = st.selectbox("Select The Report", options=[
    "MF Disbursement", "MF Sourcing", "NPA Writeoff Dashboard", "X Bucket Billing Efficiency"
])
# Clear broken cache
# win32com.client.gencache.EnsureDispatch('Excel.Application')
# Upload the report
uploaded_file = st.file_uploader("Upload the Raw Report", type=["xlsx"])


def mf_disb(uploaded_file):

    df_raw = pd.read_excel(uploaded_file, sheet_name="Branch wise", header=[0, 1, 2])
    df_raw.columns = [' | '.join([str(i) for i in col if str(i).strip() != '']).strip() for col in df_raw.columns]

    df_filtered = df_raw[df_raw["Unnamed: 2_level_0 | Unnamed: 2_level_1 | Division"]
                         .isin(['GUJARAT', 'NASHIK', 'PUNE'])].copy()

    columns_to_keep = [
        'Division', 'AREA', 'BR NAME',
        'Target | Unnamed: 6_level_1 | Count',
        'Target | Unnamed: 7_level_1 | Loan Amt (Lakhs)',
        'Disbursement - MTD | TOTAL | Count',
        'Disbursement - MTD | TOTAL | Loan Amt (Lakhs)',
        'Disbursement - MTD | TOTAL | Net Dis',
        'MTD - Potential | Count',
        'MTD - Potential | Loan Amt (Lakhs)'
    ]

    final_cols = [col for key in columns_to_keep for col in df_filtered.columns if key in col]
    df_filtered = df_filtered[final_cols]
    df_filtered.columns = [
        'Division', 'Area', 'Branch Name', 'Target Count', 'Target Loan Amt (Lakhs)',
        'Total Count', 'Total Loan Amt (Lakhs)', 'Net Disbursed',
        'MTD Potential Count', 'MTD Potential Loan Amt (Lakhs)'
    ]
    df_filtered = df_filtered.apply(lambda x: x.round(2) if x.dtype.kind in 'fc' else x)
    # Sort by Branch Name alphabetically (A to Z)
    df_filtered = df_filtered.sort_values(by="Branch Name", ascending=True)


    mf_template_path = "MF Disb Areawise.xlsx"
    wb = load_workbook(mf_template_path)
    template_sheet = wb.active

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for area, area_df in df_filtered.groupby('Area'):
        ws = wb.copy_worksheet(template_sheet)
        ws.title = str(area)[:31]
        for r_idx, row in enumerate(area_df.itertuples(index=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if isinstance(value, (int, float)) and value == 0:
                    cell.fill = red_fill
                cell.border = thin_border

        # Autofit columns based on length of data
        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=3, max_row=3 + len(area_df), max_col=12), start=1):
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.remove(template_sheet)
    if len(wb.sheetnames) > 1 and wb.sheetnames[0] == 'Sheet':
        del wb['Sheet']

    # Save for COM image capture
    # temp_excel_path = r"C:\Users\DELL\Desktop\Office Work\MF Disb\MF_Disb_Areawise_Temp.xlsx"
    # wb.save(temp_excel_path)

    # def export_excel_sheets_as_images(excel_path, output_dir):
    #     pythoncom.CoInitialize()
    #     if not os.path.exists(output_dir):
    #         os.makedirs(output_dir)

    #     excel = win32.gencache.EnsureDispatch('Excel.Application')
    #     excel.Visible = False
    #     excel.DisplayAlerts = False
    #     wb_com = excel.Workbooks.Open(excel_path)

    #     for sheet in wb_com.Sheets:
    #         try:
    #             sheet.Activate()

    #             # Determine last used row and column
    #             used_range = sheet.UsedRange
    #             last_row = used_range.Rows.Count
    #             last_col = used_range.Columns.Count
    #             last_col_letter = chr(64 + last_col) if last_col <= 26 else 'Z'

    #             cell_range = f"A1:{last_col_letter}{last_row}"
    #             sheet.Range(cell_range).CopyPicture(Format=win32.constants.xlBitmap)

    #             chart_sheet = wb_com.Sheets.Add()
    #             chart_obj = chart_sheet.ChartObjects().Add(70, 20, 600, 300)
    #             chart_obj.Chart.Paste()

    #             image_path = os.path.join(output_dir, f"{sheet.Name}.png")
    #             chart_obj.Chart.Export(Filename=image_path, FilterName="PNG")

    #             wb_com.Sheets(chart_sheet.Name).Delete()
    #         except Exception as e:
    #             print(f"❌ Failed on sheet {sheet.Name}: {e}")

    #     wb_com.Close(SaveChanges=False)
    #     excel.Quit()
    #     del excel
    #     pythoncom.CoUninitialize()

    # export_excel_sheets_as_images(
    #     excel_path=temp_excel_path,
    #     output_dir=r"C:\Users\DELL\Desktop\Office Work\MF Disb\Sheet_Images"
    # )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output



def mf_sourcing(df_raw):
    # Read file (only if needed, otherwise use df_raw directly)
    df_raw = pd.read_excel(uploaded_file, sheet_name="MF - Branch Wise Summary", header=[0, 1, 2])
    df_raw.columns = [' | '.join([str(i).strip() for i in col if str(i).strip() != '']) for col in df_raw.columns]
    df_raw = df_raw[~df_raw['Unnamed: 3_level_0 | Unnamed: 3_level_1 | AREA'].str.contains("total", case=False, na=False)]

    # Define your column names (unchanged)
    division_col = 'Unnamed: 2_level_0 | Unnamed: 2_level_1 | Division'
    area_col = 'Unnamed: 3_level_0 | Unnamed: 3_level_1 | AREA'
    branch_col = 'Unnamed: 4_level_0 | Unnamed: 4_level_1 | BR NAME'
    sale_login_count_col = 'Total Sales & Credit Login (LOS) | Sale Login | Count'
    sale_login_value_col = 'Total Sales & Credit Login (LOS) | Sale Login | Value (In Cr.)'
    rejected_count_col = "Login & Approval taken After 25th Apr'25 | Rejected | Count"
    rejected_value_col = "Login & Approval taken After 25th Apr'25 | Rejected | Value (In Cr.)"
    credit_login_count_col = "Login & Approval taken After 25th Apr'25 | Login | Count"
    credit_login_value_col = "Login & Approval taken After 25th Apr'25 | Login | Value (In Cr.)"
    approved_count_col = "Login & Approval taken After 25th Apr'25 | Approved | Count"
    approved_value_col = "Login & Approval taken After 25th Apr'25 | Approved | Value (In Cr.)"
    rfd_count_col = "Login & Approval taken After 25th Apr'25 | RFD | Count"
    rfd_value_col = "Login & Approval taken After 25th Apr'25 | RFD | Value (In Cr.)"
    disb_count_col = "Login & Approval taken After 25th Apr'25 | Disbursed | Count"
    disb_value_col = "Login & Approval taken After 25th Apr'25 | Disbursed | Value (In Cr.)"
    legal_approved_count_col = "Login & Approval taken After 25th Apr'25 | Legal Approved | Count"
    legal_approved_value_col = "Login & Approval taken After 25th Apr'25 | Legal Approved | Value (In Cr.)"

    df_raw[sale_login_count_col] -= df_raw[rejected_count_col]
    df_raw[sale_login_value_col] -= df_raw[rejected_value_col]
    df_filtered = df_raw[df_raw[division_col].isin(['GUJARAT', 'NASHIK', 'PUNE'])]

    # Keep columns
    columns_to_keep = [
        division_col, area_col, branch_col,
        sale_login_count_col, sale_login_value_col,
        credit_login_count_col, credit_login_value_col,
        approved_count_col, approved_value_col,
        rfd_count_col, rfd_value_col,
        disb_count_col, disb_value_col,
        legal_approved_count_col, legal_approved_value_col
    ]
    df_filtered = df_filtered[columns_to_keep]
    df_filtered = df_filtered.apply(lambda x: x.round(2) if x.dtype.kind in 'fc' else x)
    # Sort by Branch Name alphabetically (A to Z)
    df_filtered = df_filtered.sort_values(by=branch_col, ascending=True)

    # KPIs
    col1, col2, col3 = st.columns(3)
    col1.metric("Sale Login (Cr)", df_filtered[sale_login_value_col].sum())
    col2.metric("Login Count", int(df_filtered[credit_login_count_col].sum()))
    col3.metric("Approved (Cr)", df_filtered[approved_value_col].sum())

    col4, col5, col6 = st.columns(3)
    col4.metric("RFD Count", int(df_filtered[rfd_count_col].sum()))
    col5.metric("Disbursed (Cr)", df_filtered[disb_value_col].sum())
    col6.metric("Legal Approved Count", int(df_filtered[legal_approved_count_col].sum()))

    # Group & Chart
    group_by = st.selectbox("Group By", options=[division_col, area_col, branch_col])
    chart_data = df_filtered.groupby(group_by)[[
        sale_login_value_col, disb_value_col, credit_login_value_col,
        approved_value_col, rfd_value_col, legal_approved_value_col
    ]].sum().reset_index()
    chart_data["Potential Business"] = chart_data[credit_login_value_col] + chart_data[approved_value_col] + chart_data[rfd_value_col] + chart_data[legal_approved_value_col]
    chart_data = chart_data.rename(columns={sale_login_value_col: "Sale Login", disb_value_col: "Disbursed"})
    melted = chart_data.melt(id_vars=group_by, value_vars=["Sale Login", "Disbursed", "Potential Business"], var_name="Metric", value_name="Value (Cr.)")

    import plotly.express as px
    fig = px.bar(melted, x=group_by, y="Value (Cr.)", color="Metric", barmode="group")
    fig.update_layout(title=f"Sale Login vs Disbursed vs Potential Business by {group_by}")
    st.plotly_chart(fig, use_container_width=True)

    with st.expander("🔍 View Raw Data"):
        st.dataframe(df_filtered)

    # Excel part
    df_groupable = df_filtered[~df_filtered[area_col].str.contains("total", case=False, na=False)]
    mfs_template_path = "MF Sourcing Format.xlsx"
    wb = load_workbook(mfs_template_path)
    template_sheet = wb.active
    highlight_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for area, area_df in df_groupable.groupby(area_col):
        area_df = area_df.reset_index(drop=True)
        total_row = area_df.select_dtypes(include='number').sum()
        total_row[division_col] = ''
        total_row[area_col] = ''
        total_row[branch_col] = 'Total'
        area_df = pd.concat([area_df, pd.DataFrame([total_row])], ignore_index=True)
        ws = wb.copy_worksheet(template_sheet)
        ws.title = str(area)[:31]
        for r_idx, row in enumerate(area_df.itertuples(index=False), start=4):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
        for r_idx in range(4, ws.max_row + 1):
            if ws.cell(row=r_idx, column=3).value == 'Total':
                for c_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.fill = highlight_fill
                    cell.font = bold_font
        # adjust the column width
        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=3, max_row=3 + len(area_df), max_col=12), start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    wb.remove(template_sheet)

    # temp_excel_path = r"C:\Users\DELL\Desktop\Office Work\MF Sourcing\MF_Sourcing_Areawise_Temp.xlsx"
    # wb.save(temp_excel_path)

    # def export_excel_sheets_as_images(excel_path, output_dir):
    #     pythoncom.CoInitialize()
    #     if not os.path.exists(output_dir):
    #         os.makedirs(output_dir)
    #     excel = win32.gencache.EnsureDispatch('Excel.Application')
    #     excel.Visible = False
    #     excel.DisplayAlerts = False
    #     wb_com = excel.Workbooks.Open(excel_path)

    #     for sheet in wb_com.Sheets:
    #         try:
    #             sheet.Activate()
    #             used_range = sheet.UsedRange
    #             last_row = used_range.Rows.Count
    #             last_col = used_range.Columns.Count
    #             col_letter = chr(64 + last_col) if last_col <= 26 else 'Z'
    #             cell_range = f"A1:{col_letter}{last_row}"
    #             sheet.Range(cell_range).CopyPicture(Format=win32.constants.xlBitmap)
    #             chart_sheet = wb_com.Sheets.Add()
    #             chart_obj = chart_sheet.ChartObjects().Add(70, 20, 600, 300)
    #             chart_obj.Chart.Paste()
    #             img_path = os.path.join(output_dir, f"{sheet.Name}.png")
    #             chart_obj.Chart.Export(Filename=img_path, FilterName="PNG")
    #             wb_com.Sheets(chart_sheet.Name).Delete()
    #         except Exception as e:
    #             print(f"❌ Failed on sheet {sheet.Name}: {e}")

    #     wb_com.Close(SaveChanges=False)
    #     excel.Quit()
    #     pythoncom.CoUninitialize()

    # export_excel_sheets_as_images(
    #     excel_path=temp_excel_path,
    #     output_dir=r"C:\Users\DELL\Desktop\Office Work\MF Sourcing\Sheet_Images"
    # )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def npa_woff(df_raw):
    df_raw = pd.read_excel(uploaded_file, sheet_name="Branch wise", header=[0, 1])

    # Step 2: Extract current and previous month column names from file name
    match = re.search(r'(\d{2})-([A-Za-z]{3})-(\d{2})', uploaded_file.name)
    if match:
        day, mon, year = match.groups()
        curr_month = f"MTD -{day} {mon}'{year}"
        prev_month = pd.to_datetime(f"{day}-{mon}-{year}", format="%d-%b-%y") - pd.DateOffset(months=1)
        prev_mon_str = prev_month.strftime("MTD -%d %b'%y")
    else:
        raise ValueError("Date format not found in filename")

    # Flatten multi-level columns
    df_raw.columns = [' | '.join([str(i).strip() for i in col if str(i).strip() != '']) for col in df_raw.columns]

    # Step 3: Filter only selected divisions
    div_col = 'Unnamed: 2_level_0 | Division'
    target_divisions = ['GUJARAT', 'NASHIK', 'PUNE']
    df_filtered = df_raw[df_raw[div_col].isin(target_divisions)]

    # Column references
    area_col = 'Unnamed: 3_level_0 | AREA'
    branch_col = 'Unnamed: 5_level_0 | BR NAME'

    # Columns to keep
    columns_to_keep = [
        div_col,
        area_col,
        branch_col,
        'Fresh NPA Flow_MTD | Loans', 'Fresh NPA Flow_MTD | POS (Lakhs)',
        'Fresh Flow %_MTD | Loans', 'Fresh Flow %_MTD | POS (Lakhs)',
        'NPA_Normalization | Loans', 'NPA_Normalization | POS (Lakhs)',
        'NET NPA Variance | Loans', 'NET NPA Variance | POS (Lakhs)',
        'NET Variance (Incl WO) | Loans',
        'NET Variance (Incl WO) | POS (Lakhs)', 'Writeoff - (POS Drop) | Loans',
        'Writeoff - (POS Drop) | POS (Lakhs)', 'OTS at CPC (NPA) | Loans',
        'OTS at CPC (NPA) | POS (Lakhs)', 'OTS at CPC (Writeoff) | Loans',
        'OTS at CPC (Writeoff) | POS (Lakhs)'
    ]
    df_filtered = df_filtered[columns_to_keep]
    df_filtered = df_filtered.apply(lambda x: x.round(2) if x.dtype.kind in 'fc' else x)

    # Convert % columns from 12.34 to 0.1234 so Excel can display as proper %
    for col in ['Fresh Flow %_MTD | Loans', 'Fresh Flow %_MTD | POS (Lakhs)']:
        df_filtered[col] = df_filtered[col] / 100

    # Only group rows where area is present
    df_groupable = df_filtered[df_filtered[area_col].notna() & (df_filtered[area_col].astype(str).str.strip() != "")]
    
    # Sort by Branch Name
    df_filtered = df_filtered.sort_values(by=branch_col, ascending=True)

    # Define formatting styles
    light_pink_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")  # Light pink
    dark_pink_font = Font(color="C71585")  # Medium violet red (dark pink)

    light_green_fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")  # Light green
    dark_green_font = Font(color="006400")  # Dark green

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    highlight_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Light Yellow
    bold_font = Font(bold=True)

    format_cols = [
        'NET NPA Variance | Loans',
        'NET NPA Variance | POS (Lakhs)',
        'NET Variance (Incl WO) | Loans',
        'NET Variance (Incl WO) | POS (Lakhs)'
    ]
    percent_cols = ['Fresh Flow %_MTD | Loans', 'Fresh Flow %_MTD | POS (Lakhs)']

    # Load template workbook
    npa_template_path = 'NPA Writeoff format.xlsx'
    wb = load_workbook(npa_template_path)
    template_sheet = wb.active

    for area, area_df in df_groupable.groupby(area_col):
        area_df = area_df.reset_index(drop=True)

        # Add Area_total Row
        total_row = area_df.select_dtypes(include='number').sum(numeric_only=True)
        total_row[div_col] = ''
        total_row[area_col] = ''
        total_row[branch_col] = 'Total'
        area_df = pd.concat([area_df, pd.DataFrame([total_row])], ignore_index=True)

        # Copy template sheet
        ws = wb.copy_worksheet(template_sheet)
        ws.title = str(area)[:31]

        for r_idx, row in enumerate(area_df.itertuples(index=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border

                col_name = area_df.columns[c_idx - 1]

                # Apply % format
                if col_name in percent_cols:
                    cell.number_format = '0.00%'

                # Conditional fill based on values
                if col_name in format_cols and isinstance(value, (int, float)):
                    if value > 0:
                        cell.fill = light_pink_fill
                        cell.font = dark_pink_font
                    elif value < 0:
                        cell.fill = light_green_fill
                        cell.font = dark_green_font

        # Highlight total row
        for r_idx in range(4, ws.max_row + 1):
            if ws.cell(row=r_idx, column=3).value == 'Total':
                for c_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.fill = highlight_fill
                    cell.font = bold_font

        # Adjust column widths
        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=3, max_row=3 + len(area_df), max_col=12), start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    # Remove the template
    wb.remove(template_sheet)

    # # Save output
    # temp_excel_path = r"C:\Users\DELL\Desktop\Office Work\NPA Writeoff Dashboard\NPA_Writeoff_Dashboard_temp.xlsx"
    # wb.save(temp_excel_path)


    # def export_excel_sheets_as_images(excel_path, output_dir):
    #     pythoncom.CoInitialize()
    #     if not os.path.exists(output_dir):
    #         os.makedirs(output_dir)
    #     excel = win32.gencache.EnsureDispatch('Excel.Application')
    #     excel.Visible = False
    #     excel.DisplayAlerts = False
    #     wb_com = excel.Workbooks.Open(excel_path)

    #     for sheet in wb_com.Sheets:
    #         try:
    #             sheet.Activate()
    #             used_range = sheet.UsedRange
    #             last_row = used_range.Rows.Count
    #             last_col = used_range.Columns.Count
    #             col_letter = chr(64 + last_col) if last_col <= 26 else 'Z'
    #             cell_range = f"A1:{col_letter}{last_row}"
    #             sheet.Range(cell_range).CopyPicture(Format=win32.constants.xlBitmap)
    #             chart_sheet = wb_com.Sheets.Add()
    #             chart_obj = chart_sheet.ChartObjects().Add(100, 30, 800, 400)
    #             chart_obj.Chart.Paste()
    #             img_path = os.path.join(output_dir, f"{sheet.Name}.png")
    #             chart_obj.Chart.Export(Filename=img_path, FilterName="PNG")
    #             wb_com.Sheets(chart_sheet.Name).Delete()
    #         except Exception as e:
    #             print(f"❌ Failed on sheet {sheet.Name}: {e}")

    #     wb_com.Close(SaveChanges=False)
    #     excel.Quit()
    #     pythoncom.CoUninitialize()

    # export_excel_sheets_as_images(
    #     excel_path=temp_excel_path,
    #     output_dir=r"C:\Users\DELL\Desktop\Office Work\NPA Writeoff Dashboard\NPA_Writeoff_Dashboard_Areawise\Sheet_Images"
    # )

    # Optional: remove default 'Sheet' if exists
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb['Sheet']
    
    # Save final output
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def x_efficiency(df_raw):
    df_raw = pd.read_excel(uploaded_file, sheet_name="Branches", header=[1, 2])

    # Step 2: Extract current & previous month column names from file name
    match = re.search(r'(\d{2})-([A-Za-z]{3})-(\d{2})', uploaded_file.name)
    if match:
        day, mon, year = match.groups()
        curr_month = f"MTD -{day} {mon}'{year}"
        prev_month = pd.to_datetime(f"{day}-{mon}-{year}", format="%d-%b-%y") - pd.DateOffset(months=1)
        prev_mon_str = prev_month.strftime("MTD -%d %b'%y")
    else:
        raise ValueError("Date format not found in filename")

    # Step 3: Filter only selected divisions
    div_col = ('Val in lacs', 'DIVISION')
    target_divisions = ['GUJARAT', 'NASHIK', 'PUNE']
    df_filtered = df_raw[df_raw[div_col].isin(target_divisions)].copy()
    # Step 4: Drop any 'Variance' column (no matter what the top-level header is)
    variance_cols = [col for col in df_filtered.columns if col[1] == 'Variance']
    df_filtered.drop(columns=variance_cols, inplace=True)

    # Step 4: Efficiency drop calculation
    df_filtered[("Comparison", "Efficiency_Drop_%")] = df_filtered[(curr_month, 'Billing Eff %')] - df_filtered[(prev_mon_str, 'Billing Eff %')]

    # Step 5: Drop unwanted columns
    cols_to_drop = [
        ('Unnamed: 0_level_0', 's'),
        ('Val in lacs', 'Region'),
        ('Val in lacs', 'Cluster'),
        ('Val in lacs', 'Branch code'),
        (           "MTD - Mar'25",        'Demand AMT'),
        (           "MTD - Mar'25",       'Coll amount'),
        (           "MTD - Mar'25",   'Not coll amount'),
        (           "MTD - Mar'25",     'Billing Eff %'),
    ]
    df_filtered.drop(columns=[col for col in cols_to_drop if col in df_filtered.columns], inplace=True)

    # Step 6: Flatten columns
    def clean_col(col_tuple):
        col_str = ' | '.join(col_tuple).strip()
        return col_str.replace('Val in lacs | ', '')
    df_filtered.columns = [clean_col(col) for col in df_filtered.columns]

    # Sort by Branch Name alphabetically (A to Z)
    df_filtered = df_filtered.sort_values(by=('Branch'), ascending=True)

    # Step 7: Round values
    df_filtered = df_filtered.map(lambda x: round(x, 4) if isinstance(x, (int, float)) else x)

    # Step 8: Load format template and copy the sheet per area
    x_template_file = "Efficiency Format.xlsx"
    wb = load_workbook(x_template_file)
    template_sheet = wb.active  # the only sheet present

    area_col = 'Area'
    output_file = "Efficiency_Areawise_Sheets.xlsx"

    for area, area_df in df_filtered.groupby(area_col):
        sheet_name = str(area)[:31]

        # Copy the format sheet
        ws = wb.copy_worksheet(template_sheet)
        ws.title = sheet_name

        # Define percent format columns
        percent_cols = [col for col in area_df.columns if 'Billing Eff %' in col or 'Efficiency_Drop_%' in col]
        percent_col_indices = [area_df.columns.get_loc(col) + 1 for col in percent_cols]  # 1-based index

        # Apply borders
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Write data starting from row 3
        for r_idx, row in enumerate(area_df.itertuples(index=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Apply % formatting if column matches
                if c_idx in percent_col_indices:
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00

                # Apply border
                cell.border = thin_border

        # Apply conditional formatting for % columns
        for col_idx in percent_col_indices:
            col_letter = get_column_letter(col_idx)
            start_row = 3
            end_row = start_row + len(area_df) - 1
            data_range = f"{col_letter}{start_row}:{col_letter}{end_row}"

            rule = ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='63BE7B'
            )
            ws.conditional_formatting.add(data_range, rule)
        # adjust the column width
        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=4, max_row=6 + len(area_df), max_col=12), start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 6


    # Delete the original template sheet (only after copying all)
    if template_sheet.title in wb.sheetnames:
        del wb[template_sheet.title]
    # temp_excel_path = r"C:\Users\DELL\Desktop\Office Work\X Bucket Billing Efficiency\X_bkt_Billing_efficiency_temp.xlsx"
    # wb.save(temp_excel_path)

    # def export_excel_sheets_as_images(excel_path, output_dir):
    #     pythoncom.CoInitialize()
    #     if not os.path.exists(output_dir):
    #         os.makedirs(output_dir)
    #     excel = win32.gencache.EnsureDispatch('Excel.Application')
    #     excel.Visible = False
    #     excel.DisplayAlerts = False
    #     wb_com = excel.Workbooks.Open(excel_path)

    #     for sheet in wb_com.Sheets:
    #         try:
    #             sheet.Activate()
    #             used_range = sheet.UsedRange
    #             last_row = used_range.Rows.Count
    #             last_col = used_range.Columns.Count
    #             col_letter = chr(64 + last_col) if last_col <= 26 else 'Z'
    #             cell_range = f"A1:{col_letter}{last_row}"
    #             sheet.Range(cell_range).CopyPicture(Format=win32.constants.xlBitmap)
    #             chart_sheet = wb_com.Sheets.Add()
    #             chart_obj = chart_sheet.ChartObjects().Add(100, 30, 800, 400)
    #             chart_obj.Chart.Paste()
    #             img_path = os.path.join(output_dir, f"{sheet.Name}.png")
    #             chart_obj.Chart.Export(Filename=img_path, FilterName="PNG")
    #             wb_com.Sheets(chart_sheet.Name).Delete()
    #         except Exception as e:
    #             print(f"❌ Failed on sheet {sheet.Name}: {e}")

    #     wb_com.Close(SaveChanges=False)
    #     excel.Quit()
    #     pythoncom.CoUninitialize()

    # export_excel_sheets_as_images(
    #     excel_path=temp_excel_path,
    #     output_dir=r"C:\Users\DELL\Desktop\Office Work\X Bucket Billing Efficiency\X_bkt_Billing_efficiency_Areawise\Sheet_Images"
    # )
    

    # Save to in-memory buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- Report function mapping ---
report_handlers = {
    "MF Disbursement": {
        "func": mf_disb,
        "filename": "MF_Disb_Areawise_Final.xlsx"
    },
    "MF Sourcing": {
        "func": mf_sourcing,
        "filename": "MF_Sourcing_Areawise_Final.xlsx"
    },
    "NPA Writeoff Dashboard": {
        "func": npa_woff,
        "filename": "NPA_Writeoff_Dashboard_areawise.xlsx"
    },
    "X Bucket Billing Efficiency":{
        "func": x_efficiency,
        "filename": "X Bkt Billing Efficiency_areawise.xlsx"
    }
    # Add more reports here as you implement them
}

# --- Reusable handler block ---
if uploaded_file and report in report_handlers:
    handler = report_handlers[report]
    process_func = handler["func"]
    file_name = handler["filename"]

    output = process_func(uploaded_file)

    st.success("✅ File processed successfully!")

    st.download_button(
        label="📥 Download Final Excel File",
        data=output,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.info("🎉 File downloaded successfully once button is clicked.")




